# aio.py
import os
import csv
import json
import logging
import traceback
from .parser_tools import parser_writer
from .parser_tools.progressbar import SPACE, progress_bar
from .parser_tools.toolbox import Fieldnames

__excel_enabled = False

try:
    import openpyxl
    __excel_enabled = True
except (ImportError, ModuleNotFoundError):
    __excel_enabled = False

logger = logging.getLogger(__name__)

def path_preview(fpath):
    global __excel_enabled
    # Parse the input file
    try:
        # SARIF input - Get first result
        if fpath.endswith(('.sarif', '.json')):
            with open(fpath, "r", encoding='utf-8-sig') as read_obj:
                data = json.load(read_obj)
            p = "[ERROR] No path found"
            for run in data.get('runs', []):
                for result in run.get('results', []):
                    for location in result.get('locations', []):
                        try:
                            p = location['physicalLocation']['artifactLocation']['uri']
                            return p
                        except KeyError:
                            p = "[ERROR] No path found"
                            continue
            return p
        elif __excel_enabled:
            workbook = openpyxl.load_workbook(fpath)
            sheet = workbook[workbook.sheetnames[0]]
            
            # Extract Path header
            path_col = None
            for cell in sheet[1]:
                if cell.value == Fieldnames.PATH.value:
                    path_col = cell.column
            
            if path_col is None:
                raise ValueError("No \'Path\' column found")

            cell_preview = sheet.cell(row=2, column=path_col).value
            return cell_preview
        else:
            with open(fpath, "r", encoding='utf-8-sig') as read_obj:
                csv_reader = csv.DictReader(read_obj)
                first_row = next(csv_reader)
                cell_preview = first_row[Fieldnames.PATH.value]
                return cell_preview
    except Exception as e:
        return f"[ERROR] {e}"

def parse(fpath, scanner, substr, prepend):
    global __excel_enabled
    logger.info("Parsing %s - %s", scanner, fpath)
    
    # Keep track of row number and errors
    data = None
    row_num = 0
    total_rows = 0
    finding_count = 0
    err_count = 0
    
    # SARIF Move to different helper function
    if fpath.endswith(('.sarif', '.json')):
        finding_count, err_count = _parse_sarp_sarif(fpath, scanner, substr, prepend)
        return finding_count, err_count
    
    # Excel - Set data iterable and total_rows
    elif __excel_enabled and fpath.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(fpath)
        sheet = workbook[workbook.sheetnames[0]]
        headers = [cell.value for cell in sheet[1]]
        
        # Extract rows as dictionaries
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[col_idx]: row[col_idx] for col_idx in range(len(headers))}
            data.append(row_data)
        total_rows = len(data)
        
    # CSV - Set data iterable and total_rows
    else:
        # Get total number of findings
        with open(fpath, mode='r', encoding='utf-8-sig') as read_obj:
            total_rows = len([row[list(row.keys())[0]] for row in csv.DictReader(read_obj)])
        
        # Open csv in read
        with open(fpath, mode='r', encoding='utf-8-sig') as read_obj:
            data = csv.DictReader(read_obj)
            
    # Loop through every row in input file
    for row in data:
        try:
            row_num += 1
            progress_bar(row_num, total_rows, prefix=f'Parsing {os.path.basename(fpath)}'.rjust(SPACE))

            # Cut and prepend the paths and convert all backslashes to forward slashes
            path = str(row[Fieldnames.PATH.value]).replace(substr, "", 1)
            path = os.path.join(prepend, path).replace('\\', '/')
            
            new_row = {k:v for k,v in row.items()}
            new_row[Fieldnames.PATH.value] = path
            
            # Write row to outfile
            parser_writer.write_row(new_row)
            finding_count += 1
        except Exception:
            logger.error("Row %d of \'%s\':\n%s", row_num, fpath, traceback.format_exc())
            err_count += 1
    logger.info("Successfully processed %d findings", finding_count)
    logger.info("Number of erroneous rows: %d", err_count)
    return finding_count, err_count
# End of parse

def _parse_sarp_sarif(fpath, scanner, substr, prepend):
    # Side function to handle SARIF format
    
    result_num = 0
    total_results = 0
    finding_count = 0
    err_count = 0
    
    # Open json in read
    try:
        with open(fpath, mode='r', encoding='utf-8-sig') as f:
            data = json.load(f)
    except:
        logger.error("File \'%s\' failed to open:\n%s", fpath, traceback.format_exc())
        return finding_count, err_count + 1
    
    total_results = sum([len(run.get('results', [])) for run in data.get('runs', [])])
    
    # Each run is a scanner
    for run in data.get('runs', []):
        # Get scanner info
        t_scanner = run['tool']['driver']['name']
        
        # Iterate through results and rebuild excel column
        for result in run.get('results', []):
            result_num += 1
            progress_bar(result_num, total_results, prefix=f'Parsing {os.path.basename(fpath)}'.rjust(SPACE))
            try:
                finding_id = result.get('partialFingerprints', {}).get('findingId', '')
                new_row = {
                    Fieldnames.SCANNER.value: t_scanner,
                    Fieldnames.ID.value: finding_id
                }
                
                # Load all properties
                properties = result.get('properties', {})
                if len(properties) > 0:
                    if "cve" in properties:
                        new_row[Fieldnames.SCORING_BASIS.value] = properties["cve"]
                    elif "cwe" in properties:
                        new_row[Fieldnames.SCORING_BASIS.value] = properties["cwe"]
                        
                    field_lookup = {f.lower().replace(" ", "_"): f for f in Fieldnames.HEADERS.value}

                    for prop, val in properties.items():
                        if prop in field_lookup:
                            new_row[field_lookup[prop]] = val

                # Get path and line
                try:
                    path = result['locations'][0]['physicalLocation']['artifactLocation']['uri']
                    line = result['locations'][0]['physicalLocation']['region']['startLine']
                    
                    # Cut and prepend the paths and convert all backslashes to forward slashes
                    path = path.replace(substr, "", 1)
                    path = os.path.join(prepend, path).replace('\\', '/')
                except (KeyError, IndexError):
                    path = ''
                    line = ''
                
                new_row[Fieldnames.PATH.value] = path
                new_row[Fieldnames.LINE.value] = line

                # Type
                new_row[Fieldnames.TYPE.value] = result.get('ruleId', '')

                # Message
                new_row[Fieldnames.MESSAGE.value] = result.get('message', {}).get('text', '')
                
                # Severity
                new_row[Fieldnames.SEVERITY.value] = result.get('level', '')

                # Trace
                trace = ''
                codeflows = result.get('codeFlows', {})
                if len(codeflows) > 0:
                    try:
                        # Each location is a trace entry
                        locations = codeflows[0]['threadFlows'][0]['locations']
                        for i, location in enumerate(locations, start=1):
                            t_path = ''
                            t_line = ''
                            t_msg = ''
                            
                            # Get trace message
                            try:
                                t_msg = location['location']['message']['text']
                            except (KeyError, IndexError):
                                t_msg = ''
                            
                            # Get trace path
                            try:
                                t_path = location['location']['physicalLocation']['artifactLocation']['uri']
                                if len(t_path) > 0:
                                    # Cut and prepend the paths and convert all backslashes to forward slashes
                                    t_path = t_path.replace(substr, "", 1)
                                    t_path = os.path.join(prepend, t_path).replace('\\', '/')
                            except (KeyError, IndexError):
                                t_path = ''
                            
                            # Get trace line
                            try:
                                t_line = location['location']['physicalLocation']['region']['startLine']
                            except (KeyError, IndexError):
                                t_line = ''
                            
                            parts = [t_path, t_line, t_msg]
                            trace += f"{i}) {':'.join(str(p) for p in parts if len(str(p)) > 0)}\n"
                    except (KeyError, IndexError):
                        trace = ''
                        
                new_row[Fieldnames.TRACE.value] = trace.strip()
                
                # Final loop to fill in empty headers
                for fieldname in Fieldnames.HEADERS.value:
                    if fieldname not in new_row.keys():
                        new_row[fieldname] = ''
            
                # Write row to outfile
                parser_writer.write_row(new_row)
                finding_count += 1
            except:
                logger.error("Result ID %s of \'%s\':\n%s", finding_id, fpath, traceback.format_exc())
                err_count += 1
    return finding_count, err_count

# End of _parse_sarp_sarif
